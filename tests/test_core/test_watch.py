"""Competitive-intelligence organ — scoring math + the two integrity rules.

The rules under test are the ones that make the analysis trustworthy:

1. A missing datapoint is never a bad score (nullable score + coverage gap).
2. Evidence is append-only (corrections supersede, never overwrite).

Everything else — weighted points, alternative views, coverage, confidence
rollup — is arithmetic the scorecard and board report depend on being exact.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from core.watch import (
    WatchDimension,
    WatchManager,
    WatchScore,
    WatchSubject,
    build_scorecard,
    confidence_rollup,
    coverage_pct,
    score_subject,
    weighted_points,
)
from core.watch_seeds import get_pack, list_packs

# ── Pure math ───────────────────────────────────────────────────────────


class TestWeightedPoints:
    def test_sow_worked_example(self) -> None:
        # The SOW states it explicitly: 4/5 on a 10%-weighted dimension = 8.
        assert weighted_points(4, 10) == pytest.approx(8.0)

    def test_full_marks_yield_the_whole_weight(self) -> None:
        assert weighted_points(5, 12) == pytest.approx(12.0)

    def test_none_contributes_nothing_and_is_not_a_zero(self) -> None:
        # Critical: an unscored dimension must not behave like 0/5, which
        # would punish a brand for being opaque.
        assert weighted_points(None, 10) == 0.0
        assert weighted_points(1, 10) > weighted_points(None, 10)


class TestCoverage:
    def test_partial_coverage(self) -> None:
        subs = [{"name": "a"}, {"name": "b"}, {"name": "c"}, {"name": "d"}]
        assert coverage_pct(subs, {"a", "b"}) == pytest.approx(50.0)

    def test_full_and_empty(self) -> None:
        subs = [{"name": "a"}, {"name": "b"}]
        assert coverage_pct(subs, {"a", "b"}) == pytest.approx(100.0)
        assert coverage_pct(subs, set()) == pytest.approx(0.0)

    def test_dimension_without_subcriteria_is_one_datapoint(self) -> None:
        assert coverage_pct([], {"anything"}) == pytest.approx(100.0)
        assert coverage_pct([], set()) == pytest.approx(0.0)


class TestConfidenceRollup:
    def test_all_high(self) -> None:
        assert confidence_rollup(["high", "high"]) == "high"

    def test_mixed_is_conservative(self) -> None:
        assert confidence_rollup(["high", "low"]) == "medium"

    def test_all_low(self) -> None:
        assert confidence_rollup(["low", "low"]) == "low"

    def test_no_evidence_is_low_not_high(self) -> None:
        assert confidence_rollup([]) == "low"


def _dims() -> list[WatchDimension]:
    return [
        WatchDimension(
            dimension_id="d1",
            name="Games",
            weight_pct=60,
            view_weights={"customer_proposition": 90, "transition_priority": 10},
        ),
        WatchDimension(
            dimension_id="d2",
            name="KYC",
            weight_pct=40,
            view_weights={"customer_proposition": 10, "transition_priority": 90},
        ),
    ]


def _score(dim_id: str, score: float | None, cov: float = 100.0) -> WatchScore:
    return WatchScore(
        score_id=f"s_{dim_id}",
        subject_id="x",
        dimension_id=dim_id,
        score=score,
        coverage_pct=cov,
        confidence="high",
    )


class TestScoreSubject:
    def test_weighted_total_and_normalisation(self) -> None:
        out = score_subject(_dims(), {"d1": _score("d1", 5), "d2": _score("d2", 5)})
        assert out["raw_points"] == pytest.approx(100.0)
        assert out["normalized_pct"] == pytest.approx(100.0)
        assert out["scored_weight_pct"] == pytest.approx(100.0)

    def test_unscored_dimension_excluded_from_denominator(self) -> None:
        # Only Games (60%) scored, at 5/5. Raw is 60 — but normalised is 100,
        # because the brand is perfect *on what we actually know*. The thin
        # evidence shows up as scored_weight_pct, not as a depressed score.
        out = score_subject(_dims(), {"d1": _score("d1", 5)})
        assert out["raw_points"] == pytest.approx(60.0)
        assert out["scored_weight_pct"] == pytest.approx(60.0)
        assert out["normalized_pct"] == pytest.approx(100.0)
        assert out["unscored_dimensions"] == ["KYC"]

    def test_null_score_is_not_a_one(self) -> None:
        null_out = score_subject(
            _dims(), {"d1": _score("d1", 5), "d2": _score("d2", None)}
        )
        low_out = score_subject(_dims(), {"d1": _score("d1", 5), "d2": _score("d2", 1)})
        assert null_out["normalized_pct"] > low_out["normalized_pct"]

    def test_views_reweight_the_same_scores(self) -> None:
        scores = {"d1": _score("d1", 5), "d2": _score("d2", 1)}
        cp = score_subject(_dims(), scores, view="customer_proposition")
        tp = score_subject(_dims(), scores, view="transition_priority")
        # Strong on games, weak on KYC → looks good to a customer, poor as a
        # transition benchmark. Same scores, different question.
        assert cp["normalized_pct"] > tp["normalized_pct"]

    def test_nothing_scored_yields_none_not_zero(self) -> None:
        out = score_subject(_dims(), {})
        assert out["normalized_pct"] is None
        assert out["raw_points"] == 0.0


class TestBuildScorecard:
    def test_ranks_and_sorts_unscored_last(self) -> None:
        subjects = [
            WatchSubject(subject_id="a", name="Alpha"),
            WatchSubject(subject_id="b", name="Beta"),
            WatchSubject(subject_id="c", name="Gamma"),
        ]
        card = build_scorecard(
            subjects,
            _dims(),
            {
                "a": {"d1": _score("d1", 3), "d2": _score("d2", 3)},
                "b": {"d1": _score("d1", 5), "d2": _score("d2", 5)},
                # c has nothing
            },
        )
        names = [r["name"] for r in card["rows"]]
        assert names[0] == "Beta" and names[1] == "Alpha"
        assert names[2] == "Gamma"
        assert card["rows"][2]["rank"] is None  # unscored is unranked, not last-place
        assert card["weights_valid"] is True

    def test_flags_invalid_weight_total(self) -> None:
        dims = [WatchDimension(dimension_id="d1", name="Only", weight_pct=70)]
        card = build_scorecard([WatchSubject(subject_id="a", name="A")], dims, {})
        assert card["weights_valid"] is False
        assert card["weight_total_pct"] == pytest.approx(70.0)


# ── Seed pack integrity ─────────────────────────────────────────────────


class TestSeedPack:
    def test_pack_is_listed(self) -> None:
        assert any(p["name"] == "social_casino_t1" for p in list_packs())

    def test_sow_shape(self) -> None:
        pack = get_pack("social_casino_t1")
        assert pack is not None
        assert len(pack["dimensions"]) == 12
        assert len(pack["subjects"]) == 14

    def test_every_weight_vector_sums_to_100(self) -> None:
        dims = get_pack("social_casino_t1")["dimensions"]
        assert sum(d["weight_pct"] for d in dims) == pytest.approx(100.0)
        for view in ("customer_proposition", "transition_priority"):
            total = sum(d["view_weights"][view] for d in dims)
            assert total == pytest.approx(100.0), f"{view} sums to {total}"

    def test_subcriteria_sum_to_100_each(self) -> None:
        for d in get_pack("social_casino_t1")["dimensions"]:
            total = sum(s["weight_pct"] for s in d["subcriteria"])
            assert total == pytest.approx(100.0), f"{d['name']} sums to {total}"

    def test_cadences_follow_the_sow_refresh_rule(self) -> None:
        dims = {
            d["name"]: d["refresh_cadence"]
            for d in get_pack("social_casino_t1")["dimensions"]
        }
        assert dims["Promotional proposition and generosity"] == "weekly"
        assert dims["Marketing proposition"] == "weekly"
        assert dims["Available P&L, accounts and financial strength"] == "quarterly"
        assert dims["State availability and variation"] == "quarterly"

    def test_our_own_brands_are_flagged(self) -> None:
        subs = get_pack("social_casino_t1")["subjects"]
        assert {s["name"] for s in subs if s.get("is_self")} == {"Pulsz", "Pulsz Bingo"}


# ── Integrity rules against a real DB ───────────────────────────────────


@pytest.fixture
async def wm(tmp_path):
    from core.database import Database

    db = Database(str(tmp_path / "watch.db"))
    await db.initialize()
    yield WatchManager(db)
    await db.close()


async def _seeded(wm: WatchManager, company: str = "c1"):
    subj = await wm.add_subject(
        name="Rival", company_id=company, url="https://r.example"
    )
    dim = await wm.upsert_dimension(
        name="Promos",
        company_id=company,
        weight_pct=100,
        subcriteria=[
            {"name": "welcome", "weight_pct": 50},
            {"name": "ongoing", "weight_pct": 50},
        ],
    )
    return subj, dim


class TestNoEvidenceNoScore:
    @pytest.mark.asyncio
    async def test_scoring_without_evidence_is_refused(self, wm) -> None:
        subj, dim = await _seeded(wm)
        with pytest.raises(ValueError, match="no evidence"):
            await wm.set_score(
                company_id="c1",
                subject_id=subj.subject_id,
                dimension_id=dim.dimension_id,
                score=4,
            )

    @pytest.mark.asyncio
    async def test_null_score_records_the_gap_instead(self, wm) -> None:
        subj, dim = await _seeded(wm)
        s = await wm.set_score(
            company_id="c1",
            subject_id=subj.subject_id,
            dimension_id=dim.dimension_id,
            score=None,
            rationale="opaque",
        )
        assert s.score is None
        assert s.coverage_pct == pytest.approx(0.0)

    @pytest.mark.asyncio
    async def test_coverage_reflects_evidence_breadth(self, wm) -> None:
        subj, dim = await _seeded(wm)
        await wm.add_evidence(
            company_id="c1",
            subject_id=subj.subject_id,
            dimension_id=dim.dimension_id,
            subcriterion="welcome",
            claim="5 SC on signup",
            confidence="high",
        )
        s = await wm.set_score(
            company_id="c1",
            subject_id=subj.subject_id,
            dimension_id=dim.dimension_id,
            score=4,
        )
        # One of two sub-criteria evidenced.
        assert s.coverage_pct == pytest.approx(50.0)
        assert s.confidence == "high"
        assert len(s.evidence_ids) == 1

    @pytest.mark.asyncio
    async def test_out_of_range_score_rejected(self, wm) -> None:
        subj, dim = await _seeded(wm)
        with pytest.raises(ValueError, match="between 1 and 5"):
            await wm.set_score(
                company_id="c1",
                subject_id=subj.subject_id,
                dimension_id=dim.dimension_id,
                score=9,
            )


class TestEvidenceIsAppendOnly:
    @pytest.mark.asyncio
    async def test_supersede_retains_history(self, wm) -> None:
        subj, dim = await _seeded(wm)
        old = await wm.add_evidence(
            company_id="c1",
            subject_id=subj.subject_id,
            dimension_id=dim.dimension_id,
            claim="min purchase $9.99",
        )
        await wm.add_evidence(
            company_id="c1",
            subject_id=subj.subject_id,
            dimension_id=dim.dimension_id,
            claim="min purchase $4.99",
            supersedes=old.evidence_id,
        )
        live = await wm.list_evidence("c1", subject_id=subj.subject_id)
        history = await wm.list_evidence(
            "c1", subject_id=subj.subject_id, include_superseded=True
        )
        assert [e.claim for e in live] == ["min purchase $4.99"]
        assert len(history) == 2  # the old observation survives for diffing

    @pytest.mark.asyncio
    async def test_invalid_enums_rejected(self, wm) -> None:
        subj, dim = await _seeded(wm)
        for kwargs, match in (
            ({"confidence": "certain"}, "confidence"),
            ({"customer_state": "whale"}, "customer_state"),
            ({"collector": "robot"}, "collector"),
        ):
            with pytest.raises(ValueError, match=match):
                await wm.add_evidence(
                    company_id="c1",
                    subject_id=subj.subject_id,
                    dimension_id=dim.dimension_id,
                    claim="x",
                    **kwargs,
                )

    @pytest.mark.asyncio
    async def test_human_collected_evidence_is_labelled(self, wm) -> None:
        # Authenticated states are operator-collected by policy; the register
        # must record who observed it so the provenance is auditable.
        subj, dim = await _seeded(wm)
        ev = await wm.add_evidence(
            company_id="c1",
            subject_id=subj.subject_id,
            dimension_id=dim.dimension_id,
            claim="VIP host assigned at $500",
            customer_state="vip",
            collector="human",
        )
        assert ev.collector == "human" and ev.customer_state == "vip"


class TestCompanyScoping:
    @pytest.mark.asyncio
    async def test_subjects_do_not_leak_across_companies(self, wm) -> None:
        await wm.add_subject(name="Rival", company_id="c1")
        await wm.add_subject(name="Other", company_id="c2")
        assert [s.name for s in await wm.list_subjects("c1")] == ["Rival"]
        assert [s.name for s in await wm.list_subjects("c2")] == ["Other"]

    @pytest.mark.asyncio
    async def test_archived_subject_hidden_but_retained(self, wm) -> None:
        subj, _ = await _seeded(wm)
        await wm.archive_subject(subj.subject_id)
        assert await wm.list_subjects("c1") == []
        assert len(await wm.list_subjects("c1", include_archived=True)) == 1


# ── P2: change detection, staleness, exports ────────────────────────────


def _card(rows: list[dict], generated_at: str = "2026-01-01T00:00:00+00:00") -> dict:
    """Minimal scorecard payload shaped like build_scorecard's output."""
    return {
        "generated_at": generated_at,
        "dimensions": [{"name": "Promos", "weight_pct": 100, "cadence": "monthly"}],
        "weight_total_pct": 100,
        "weights_valid": True,
        "rows": rows,
    }


def _row(name: str, score, rank=1, cov=100.0, group="G") -> dict:
    return {
        "subject_id": name.lower(),
        "name": name,
        "group": group,
        "is_self": False,
        "rank": rank,
        "overall": {"normalized_pct": None if score is None else score * 20},
        "views": {},
        "dimensions": {
            "Promos": {"score": score, "coverage_pct": cov, "confidence": "high"}
        },
    }


class TestDiffScorecards:
    def test_material_score_move_detected(self) -> None:
        from core.watch import diff_scorecards

        d = diff_scorecards(_card([_row("A", 3)]), _card([_row("A", 5)]))
        kinds = [i["kind"] for c in d["changed"] for i in c["items"]]
        assert "score" in kinds
        assert d["material_count"] >= 1

    def test_sub_threshold_move_is_not_material(self) -> None:
        from core.watch import diff_scorecards

        # A monthly report that surfaces every wobble gets ignored.
        d = diff_scorecards(
            _card([_row("A", 3)]), _card([_row("A", 3.5)]), min_score_delta=1.0
        )
        assert d["changed"] == []
        assert d["material_count"] == 0

    def test_newly_scored_is_always_material(self) -> None:
        from core.watch import diff_scorecards

        d = diff_scorecards(_card([_row("A", None)]), _card([_row("A", 2)]))
        kinds = [i["kind"] for c in d["changed"] for i in c["items"]]
        assert "newly_scored" in kinds  # new knowledge, never noise

    def test_score_withdrawn_is_material(self) -> None:
        from core.watch import diff_scorecards

        d = diff_scorecards(_card([_row("A", 4)]), _card([_row("A", None)]))
        kinds = [i["kind"] for c in d["changed"] for i in c["items"]]
        assert "score_withdrawn" in kinds

    def test_rank_change_detected(self) -> None:
        from core.watch import diff_scorecards

        d = diff_scorecards(
            _card([_row("A", 3, rank=1), _row("B", 2, rank=2)]),
            _card([_row("A", 3, rank=2), _row("B", 5, rank=1)]),
        )
        kinds = [i["kind"] for c in d["changed"] for i in c["items"]]
        assert "rank" in kinds

    def test_coverage_shift_detected(self) -> None:
        from core.watch import diff_scorecards

        d = diff_scorecards(
            _card([_row("A", 3, cov=20.0)]), _card([_row("A", 3, cov=90.0)])
        )
        kinds = [i["kind"] for c in d["changed"] for i in c["items"]]
        assert "coverage" in kinds

    def test_entrants_and_exits(self) -> None:
        from core.watch import diff_scorecards

        d = diff_scorecards(_card([_row("A", 3)]), _card([_row("A", 3), _row("B", 4)]))
        assert d["added_subjects"] == ["B"]
        d2 = diff_scorecards(_card([_row("A", 3), _row("B", 4)]), _card([_row("A", 3)]))
        assert d2["removed_subjects"] == ["B"]

    def test_identical_scorecards_are_quiet(self) -> None:
        from core.watch import diff_scorecards

        rows = [_row("A", 3), _row("B", 4, rank=2)]
        d = diff_scorecards(_card(rows), _card(rows))
        assert d["material_count"] == 0


class TestSnapshotsAndStaleness:
    @pytest.mark.asyncio
    async def test_snapshot_roundtrip_and_diff(self, wm) -> None:
        subj, dim = await _seeded(wm)
        await wm.add_evidence(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, subcriterion="welcome", claim="x",
        )
        await wm.set_score(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, score=2,
        )
        snap = await wm.take_snapshot("c1", label="baseline")
        assert snap
        assert (await wm.list_snapshots("c1"))[0]["label"] == "baseline"

        await wm.set_score(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, score=5,
        )
        diff = await wm.diff_since_snapshot("c1")
        assert diff is not None
        assert diff["against_snapshot"] == snap
        assert diff["material_count"] >= 1

    @pytest.mark.asyncio
    async def test_diff_without_snapshot_returns_none(self, wm) -> None:
        await _seeded(wm)
        assert await wm.diff_since_snapshot("c1") is None

    @pytest.mark.asyncio
    async def test_never_observed_pairs_are_reported(self, wm) -> None:
        await _seeded(wm)
        gaps = await wm.staleness("c1")
        assert len(gaps) == 1
        assert gaps[0]["status"] == "never_observed"
        assert gaps[0]["last_observed"] is None

    @pytest.mark.asyncio
    async def test_fresh_evidence_is_not_stale(self, wm) -> None:
        subj, dim = await _seeded(wm)
        await wm.add_evidence(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, claim="just observed",
        )
        assert await wm.staleness("c1") == []

    @pytest.mark.asyncio
    async def test_evidence_older_than_cadence_is_stale(self, wm) -> None:
        from datetime import UTC, datetime, timedelta

        subj, dim = await _seeded(wm)  # dimension defaults to monthly (30d)
        old = (datetime.now(UTC) - timedelta(days=90)).isoformat()
        await wm.add_evidence(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, claim="ancient", observed_at=old,
        )
        gaps = await wm.staleness("c1")
        assert len(gaps) == 1
        assert gaps[0]["status"] == "stale"
        assert gaps[0]["age_days"] >= 89

    @pytest.mark.asyncio
    async def test_evidence_export_resolves_names(self, wm) -> None:
        subj, dim = await _seeded(wm)
        await wm.add_evidence(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, claim="c", source_url="https://e.example",
        )
        rows = await wm.evidence_with_names("c1")
        assert rows[0]["subject"] == "Rival" and rows[0]["dimension"] == "Promos"


class TestXlsxExport:
    @pytest.mark.asyncio
    async def test_workbook_has_four_sheets_and_blanks_not_zeros(
        self, wm, tmp_path
    ) -> None:
        from openpyxl import load_workbook

        from core.watch_xlsx import render_scorecard_xlsx

        scored = await wm.add_subject(name="Scored", company_id="c1")
        await wm.add_subject(name="Unscored", company_id="c1")
        dim = await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        await wm.add_evidence(
            company_id="c1", subject_id=scored.subject_id,
            dimension_id=dim.dimension_id, claim="x",
        )
        await wm.set_score(
            company_id="c1", subject_id=scored.subject_id,
            dimension_id=dim.dimension_id, score=4,
        )

        out = tmp_path / "card.xlsx"
        path = render_scorecard_xlsx(
            await wm.scorecard("c1"),
            dimensions=await wm.list_dimensions("c1"),
            evidence=await wm.evidence_with_names("c1"),
            staleness=await wm.staleness("c1"),
            path=out,
        )
        wb = load_workbook(path)
        assert wb.sheetnames == ["Scorecard", "Weights", "Evidence", "Gaps"]

        ws = wb["Scorecard"]
        header = [c.value for c in ws[4]]
        col = header.index("Promos") + 1
        cells = {
            r[1].value: r[col - 1].value
            for r in ws.iter_rows(min_row=5, max_row=ws.max_row)
            if r[1].value in ("Scored", "Unscored")
        }
        assert cells["Scored"] == 4
        # The whole point: absent evidence renders blank, never 0.
        assert cells["Unscored"] is None


class TestBoardReportTool:
    """The report must be useful on day one and honest when no model is
    available — it ships facts either way."""

    async def _tool(self, wm):
        from tools.watch.tools import WatchBoardReportTool

        t = WatchBoardReportTool()
        t._watch_manager = wm
        return t

    @pytest.mark.asyncio
    async def test_first_cycle_states_the_baseline(self, wm) -> None:
        await _seeded(wm)
        t = await self._tool(wm)
        res = await t.execute({"company_id": "c1", "take_snapshot": False})
        assert res.success
        assert "First reporting cycle" in res.data["markdown"]
        assert res.data["material_count"] == 0

    @pytest.mark.asyncio
    async def test_without_router_it_ships_facts_and_says_so(self, wm) -> None:
        subj, dim = await _seeded(wm)
        await wm.add_evidence(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, claim="x",
        )
        await wm.set_score(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, score=2,
        )
        await wm.take_snapshot("c1", label="base")
        await wm.set_score(
            company_id="c1", subject_id=subj.subject_id,
            dimension_id=dim.dimension_id, score=5,
        )
        t = await self._tool(wm)  # no router injected
        res = await t.execute({"company_id": "c1", "take_snapshot": False})
        md = res.data["markdown"]
        assert "Material changes" in md and "2.0/5 → 5.0/5" in md
        assert res.data["judged_items"] == 0
        assert "Implications not generated" in md

    @pytest.mark.asyncio
    async def test_gaps_are_framed_as_absence_not_weakness(self, wm) -> None:
        await _seeded(wm)
        t = await self._tool(wm)
        md = (await t.execute({"company_id": "c1", "take_snapshot": False})).data[
            "markdown"
        ]
        assert "never observed" in md
        assert "not weaknesses" in md

    @pytest.mark.asyncio
    async def test_snapshot_taken_by_default(self, wm) -> None:
        await _seeded(wm)
        t = await self._tool(wm)
        res = await t.execute({"company_id": "c1"})
        assert res.data["snapshot_id"]
        assert len(await wm.list_snapshots("c1")) == 1


# ── P3: autonomous collection, proxy pool, refresh queue ────────────────


_PAGE = (
    "<html><head><style>.a{color:red}</style><script>var x=1;</script></head>"
    "<body><h1>Welcome</h1>"
    "<p>Get 57,500 Gold Coins &amp; 27.5 Sweepstakes Coins on your first purchase.</p>"
    "<p>Minimum purchase is $4.99 per package.</p></body></html>"
)


class TestHtmlToText:
    def test_drops_script_and_style_bodies(self) -> None:
        from core.watch_observe import html_to_text

        out = html_to_text(_PAGE)
        assert "var x=1" not in out and "color:red" not in out

    def test_unescapes_entities_and_collapses_whitespace(self) -> None:
        from core.watch_observe import html_to_text

        out = html_to_text(_PAGE)
        assert "Gold Coins & 27.5" in out
        assert "  " not in out

    def test_empty_input_is_safe(self) -> None:
        from core.watch_observe import html_to_text

        assert html_to_text("") == ""


class TestExcerptVerification:
    """The anti-hallucination guarantee: a claim survives only if its quote is
    genuinely in the source."""

    def test_real_quote_verifies_despite_case_and_spacing(self) -> None:
        from core.watch_observe import html_to_text, verify_excerpt

        text = html_to_text(_PAGE)
        assert verify_excerpt("get 57,500 gold coins &  27.5 sweepstakes coins", text)

    def test_fabricated_quote_is_rejected(self) -> None:
        from core.watch_observe import html_to_text, verify_excerpt

        text = html_to_text(_PAGE)
        assert not verify_excerpt(
            "Claim your daily login bonus of 10,000 Gold Coins every day.", text
        )

    def test_trivially_short_quote_is_rejected(self) -> None:
        from core.watch_observe import html_to_text, verify_excerpt

        # A 7-char quote would match half the web and prove nothing.
        assert not verify_excerpt("Welcome", html_to_text(_PAGE))

    def test_filter_splits_and_explains_rejections(self) -> None:
        from core.watch_observe import filter_verified_claims, html_to_text

        text = html_to_text(_PAGE)
        claims = [
            {"claim": "real", "excerpt": "Minimum purchase is $4.99 per package."},
            {"claim": "made up", "excerpt": "We offer a 200% cashback guarantee always"},
            {"claim": "too short", "excerpt": "Welcome"},
            {"claim": "", "excerpt": "Minimum purchase is $4.99 per package."},
        ]
        ok, bad = filter_verified_claims(claims, text)
        assert [c["claim"] for c in ok] == ["real"]
        assert {b["reason"] for b in bad} == {
            "excerpt not found in source",
            "excerpt too short",
        }


class _FakeRouter:
    """Returns one verifiable claim and one fabrication."""

    def __init__(self) -> None:
        self.calls = 0

    async def complete(self, **kwargs):
        import json as _json
        from types import SimpleNamespace

        self.calls += 1
        return SimpleNamespace(
            content=_json.dumps(
                {
                    "claims": [
                        {
                            "subcriterion": "welcome",
                            "claim": "First purchase grants 57,500 GC and 27.5 SC",
                            "value_text": "57,500 GC",
                            "excerpt": (
                                "Get 57,500 Gold Coins & 27.5 Sweepstakes Coins "
                                "on your first purchase."
                            ),
                        },
                        {
                            "subcriterion": "ongoing",
                            "claim": "Daily 10,000 GC login bonus",
                            "value_text": "10,000 GC",
                            "excerpt": "Claim your daily login bonus of 10,000 GC daily.",
                        },
                    ]
                }
            )
        )


class TestWatchObserveTool:
    async def _tool(self, wm, router=None):
        from tools.watch.tools import WatchObserveTool

        t = WatchObserveTool()
        t._watch_manager = wm
        t._router = router if router is not None else _FakeRouter()
        t._config = None
        return t

    @pytest.mark.asyncio
    async def test_only_verifiable_claims_are_persisted(
        self, wm, monkeypatch
    ) -> None:
        import core.watch_observe as wo

        subj = await wm.add_subject(
            name="Rival", company_id="c1", url="https://rival.example"
        )
        await wm.upsert_dimension(
            name="Promos", company_id="c1", weight_pct=100,
            subcriteria=[{"name": "welcome", "weight_pct": 100}],
        )
        text = wo.html_to_text(_PAGE)

        async def fake_fetch(url, **kw):
            return (text, None)

        monkeypatch.setattr(wo, "fetch_page", fake_fetch)
        res = await (await self._tool(wm)).execute(
            {"subject": "Rival", "dimension": "Promos", "company_id": "c1"}
        )
        assert res.success
        assert res.data["evidence_written"] == 1
        assert res.data["claims_rejected"] == 1

        rows = await wm.list_evidence("c1", subject_id=subj.subject_id)
        assert len(rows) == 1
        assert "57,500" in rows[0].claim
        # Agent collection is public/logged-out and labelled as such.
        assert rows[0].collector == "agent"
        assert rows[0].customer_state == "logged_out"
        assert rows[0].source_url == "https://rival.example"

    @pytest.mark.asyncio
    async def test_fetch_failure_is_reported_not_fabricated(
        self, wm, monkeypatch
    ) -> None:
        import core.watch_observe as wo

        await wm.add_subject(name="Rival", company_id="c1", url="https://x.example")
        await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)

        async def dead_fetch(url, **kw):
            return ("", "HTTP 403")

        monkeypatch.setattr(wo, "fetch_page", dead_fetch)
        res = await (await self._tool(wm)).execute(
            {"subject": "Rival", "dimension": "Promos", "company_id": "c1"}
        )
        assert res.success and res.data["evidence_written"] == 0
        assert res.data["pages"][0]["error"] == "HTTP 403"
        assert await wm.list_evidence("c1") == []

    @pytest.mark.asyncio
    async def test_requires_a_router(self, wm) -> None:
        await wm.add_subject(name="Rival", company_id="c1", url="https://x.example")
        await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        t = await self._tool(wm)
        t._router = None
        res = await t.execute(
            {"subject": "Rival", "dimension": "Promos", "company_id": "c1"}
        )
        assert not res.success and "router" in res.error

    @pytest.mark.asyncio
    async def test_missing_url_is_a_clear_error(self, wm) -> None:
        await wm.add_subject(name="NoUrl", company_id="c1")
        await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        res = await (await self._tool(wm)).execute(
            {"subject": "NoUrl", "dimension": "Promos", "company_id": "c1"}
        )
        assert not res.success and "no URL" in res.error


class TestProxyPool:
    def test_state_exit_selected_from_pool(self) -> None:
        from core.config import ProxyConfig

        p = ProxyConfig(
            enabled=True, type="http", host="default.example", port=1,
            pool=[
                {"state": "TX", "host": "tx.example", "port": 8080,
                 "username": "u", "password": "p"},
                {"state": "CA", "host": "ca.example", "port": 8081},
            ],
        )
        assert p.exit_for_state("tx")["host"] == "tx.example"
        assert p.request_proxy_url("TX") == "http://u:p@tx.example:8080"
        assert p.request_proxy_url("CA") == "http://ca.example:8081"

    def test_falls_back_to_single_exit(self) -> None:
        from core.config import ProxyConfig

        p = ProxyConfig(enabled=True, type="http", host="d.example", port=3128)
        assert p.exit_for_state("TX") is None
        assert p.request_proxy_url("TX") == "http://d.example:3128"
        assert p.request_proxy_url("n/a") == "http://d.example:3128"

    def test_disabled_proxy_means_direct(self) -> None:
        from core.config import ProxyConfig

        assert ProxyConfig(enabled=False, host="d.example", port=1).request_proxy_url(
            "TX"
        ) == ""

    def test_pool_works_even_when_single_proxy_disabled(self) -> None:
        from core.config import ProxyConfig

        # A geo pool is about *observation*, not about routing all traffic.
        p = ProxyConfig(
            enabled=False,
            pool=[{"state": "NY", "host": "ny.example", "port": 9, "type": "http"}],
        )
        assert p.request_proxy_url("NY") == "http://ny.example:9"


class TestWatchQueueTool:
    async def _tool(self, wm):
        from tools.watch.tools import WatchQueueTool

        t = WatchQueueTool()
        t._watch_manager = wm
        return t

    @pytest.mark.asyncio
    async def test_never_observed_ranks_above_merely_stale(self, wm) -> None:
        from datetime import UTC, datetime, timedelta

        s1 = await wm.add_subject(name="Stale", company_id="c1")
        await wm.add_subject(name="Never", company_id="c1")
        dim = await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        await wm.add_evidence(
            company_id="c1", subject_id=s1.subject_id, dimension_id=dim.dimension_id,
            claim="old",
            observed_at=(datetime.now(UTC) - timedelta(days=99)).isoformat(),
        )
        res = await (await self._tool(wm)).execute({"company_id": "c1"})
        assert res.data["never_observed"] == 1 and res.data["stale"] == 1
        assert res.data["queue"][0]["subject"] == "Never"

    @pytest.mark.asyncio
    async def test_cadence_filter(self, wm) -> None:
        await wm.add_subject(name="A", company_id="c1")
        await wm.upsert_dimension(
            name="Weekly thing", company_id="c1", refresh_cadence="weekly"
        )
        await wm.upsert_dimension(
            name="Quarterly thing", company_id="c1", refresh_cadence="quarterly"
        )
        res = await (await self._tool(wm)).execute(
            {"company_id": "c1", "cadence": "weekly"}
        )
        assert res.data["due_count"] == 1
        assert res.data["queue"][0]["dimension"] == "Weekly thing"

    @pytest.mark.asyncio
    async def test_schedule_requires_scheduler(self, wm) -> None:
        res = await (await self._tool(wm)).execute(
            {"company_id": "c1", "action": "schedule"}
        )
        assert not res.success and "scheduler" in res.error


# ── Browser escalation + one-command analysis ───────────────────────────


class _FakeBrowser:
    """Stands in for the real Chrome bridge."""

    def __init__(self, text: str = "", html: str = "", fail: bool = False) -> None:
        self.text, self.html, self.fail = text, html, fail
        self.calls: list[str] = []

    async def call_tool(self, name: str, params: dict):
        self.calls.append(name)
        if self.fail:
            raise RuntimeError("chrome unavailable")
        if name == "browser_navigate":
            return {"url": params.get("url"), "title": "t"}
        if name == "browser_extract":
            return {"text": self.text}
        if name == "browser_get_html":
            return {"html": self.html}
        return {}


class TestBrowserEscalation:
    """A JS shell must not defeat an agent that has a real browser."""

    @pytest.mark.asyncio
    async def test_js_shell_escalates_to_browser(self, monkeypatch) -> None:
        import core.watch_observe as wo

        rendered = "Rendered by Chrome. " + ("Real offer content. " * 60)

        async def shell_fetch(url, **kw):
            return ("App Loading", None)  # what a plain client sees

        monkeypatch.setattr(wo, "fetch_page", shell_fetch)
        browser = _FakeBrowser(text=rendered)
        text, err, method = await wo.fetch_page_best_effort(
            "https://spa.example", browser_manager=browser
        )
        assert err is None
        assert method == "browser"
        assert "Rendered by Chrome" in text
        assert browser.calls[:2] == ["browser_navigate", "browser_extract"]

    @pytest.mark.asyncio
    async def test_blocked_request_escalates_to_browser(self, monkeypatch) -> None:
        import core.watch_observe as wo

        async def blocked(url, **kw):
            return ("", "HTTP 403")

        monkeypatch.setattr(wo, "fetch_page", blocked)
        text, err, method = await wo.fetch_page_best_effort(
            "https://blocked.example",
            browser_manager=_FakeBrowser(text="Content behind the block. " * 30),
        )
        assert err is None and method == "browser"
        assert "behind the block" in text

    @pytest.mark.asyncio
    async def test_rich_page_never_wakes_the_browser(self, monkeypatch) -> None:
        import core.watch_observe as wo

        async def rich(url, **kw):
            return ("Plenty of real content here. " * 40, None)

        monkeypatch.setattr(wo, "fetch_page", rich)
        browser = _FakeBrowser(text="should not be used")
        text, err, method = await wo.fetch_page_best_effort(
            "https://static.example", browser_manager=browser
        )
        # The browser is slow and shared — don't spend it when HTTP sufficed.
        assert method == "http" and browser.calls == []

    @pytest.mark.asyncio
    async def test_browser_falls_back_to_raw_html(self, monkeypatch) -> None:
        import core.watch_observe as wo

        async def shell(url, **kw):
            return ("", "HTTP 403")

        monkeypatch.setattr(wo, "fetch_page", shell)
        browser = _FakeBrowser(
            text="", html="<body><p>" + ("Recovered via HTML. " * 30) + "</p></body>"
        )
        text, err, method = await wo.fetch_page_best_effort(
            "https://x.example", browser_manager=browser
        )
        assert err is None and "Recovered via HTML" in text
        assert "browser_get_html" in browser.calls

    @pytest.mark.asyncio
    async def test_thin_page_still_read_when_no_browser(self, monkeypatch) -> None:
        import core.watch_observe as wo

        async def sparse(url, **kw):
            return ("Short but genuine page text.", None)

        monkeypatch.setattr(wo, "fetch_page", sparse)
        text, err, method = await wo.fetch_page_best_effort("https://s.example")
        # Sparse is not the same as unreadable — don't throw away real evidence.
        assert err is None and method == "http_thin" and text


class TestLinkDiscovery:
    def test_finds_commercially_relevant_subpages(self) -> None:
        from core.watch_observe import discover_links

        html = """
        <a href="/terms-and-conditions">Terms</a>
        <a href="/promotions">Promotions</a>
        <a href="https://other.example/promo">Offsite promo</a>
        <a href="/about-us">About</a>
        <a href="#top">Top</a>
        <a href="/help/payment-methods">Payments</a>
        """
        links = discover_links(html, "https://brand.example")
        assert "https://brand.example/terms-and-conditions" in links
        assert "https://brand.example/promotions" in links
        assert "https://brand.example/help/payment-methods" in links
        # Same-site only, and uninteresting pages are skipped.
        assert not any("other.example" in link for link in links)
        assert not any("about-us" in link for link in links)

    def test_handles_junk_input(self) -> None:
        from core.watch_observe import discover_links

        assert discover_links("", "https://b.example") == []
        assert discover_links("<a href='mailto:x@y.z'>terms</a>", "https://b.example") == []


class _AnalyzeRouter:
    """Extracts one good + one fabricated claim, then scores."""

    async def complete(self, messages, **kwargs):
        import json as _json
        from types import SimpleNamespace

        system = messages[0]["content"]
        if "You score one brand" in system:
            return SimpleNamespace(
                content=_json.dumps(
                    {"score": 4, "rationale": "strong welcome offer", "provisional": True}
                )
            )
        return SimpleNamespace(
            content=_json.dumps(
                {
                    "claims": [
                        {
                            "dimension": "Promos",
                            "subcriterion": "welcome",
                            "claim": "57,500 GC on first purchase",
                            "value_text": "57,500 GC",
                            "excerpt": "Get 57,500 Gold Coins on your first purchase today.",
                        },
                        {
                            "dimension": "Promos",
                            "subcriterion": "ongoing",
                            "claim": "Daily bonus of 1m GC",
                            "value_text": "1m GC",
                            "excerpt": "We hand out one million Gold Coins every single day.",
                        },
                    ]
                }
            )
        )


class TestWatchAnalyzeTool:
    """The one-command path: read → verify → score → save."""

    async def _tool(self, wm):
        from tools.watch.tools import WatchAnalyzeTool

        t = WatchAnalyzeTool()
        t._watch_manager = wm
        t._router = _AnalyzeRouter()
        t._config = None
        t._browser_manager = None
        return t

    @pytest.mark.asyncio
    async def test_end_to_end_writes_evidence_scores_and_files(
        self, wm, tmp_path, monkeypatch
    ) -> None:
        import core.watch_observe as wo

        page = (
            "Get 57,500 Gold Coins on your first purchase today. "
            + ("Play hundreds of games. " * 40)
        )

        async def fake_collect(start_url, **kw):
            return [{"url": start_url, "text": page, "error": None, "method": "http"}]

        monkeypatch.setattr(wo, "collect_pages", fake_collect)

        await wm.add_subject(
            name="Rival", company_id="c1", url="https://rival.example"
        )
        await wm.upsert_dimension(
            name="Promos", company_id="c1", weight_pct=100,
            subcriteria=[{"name": "welcome", "weight_pct": 50},
                         {"name": "ongoing", "weight_pct": 50}],
        )

        res = await (await self._tool(wm)).execute(
            {"subject": "Rival", "company_id": "c1", "out_dir": str(tmp_path)}
        )
        assert res.success, res.error
        # Only the claim whose quote is on the page survives.
        assert res.data["evidence_written"] == 1
        assert res.data["claims_rejected"] == 1
        assert res.data["dimensions_scored"] == 1
        assert res.data["scores"][0]["score"] == 4
        assert res.data["scores"][0]["provisional"] is True

        saved = res.data["saved"]
        assert Path(saved["scorecard"]).exists()
        assert Path(saved["report"]).exists()

    @pytest.mark.asyncio
    async def test_unknown_brand_lists_the_tracked_ones(self, wm) -> None:
        await wm.add_subject(name="Known", company_id="c1")
        await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        res = await (await self._tool(wm)).execute(
            {"subject": "Nope", "company_id": "c1", "save": False}
        )
        assert not res.success
        assert "not tracked" in res.error and "Known" in res.error

    @pytest.mark.asyncio
    async def test_untracked_brand_with_url_is_added(self, wm, monkeypatch) -> None:
        import core.watch_observe as wo

        async def fake_collect(start_url, **kw):
            return [{"url": start_url, "text": "x" * 900, "error": None, "method": "http"}]

        monkeypatch.setattr(wo, "collect_pages", fake_collect)
        await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        res = await (await self._tool(wm)).execute(
            {
                "subject": "Fresh",
                "url": "https://fresh.example",
                "company_id": "c1",
                "save": False,
            }
        )
        assert res.success
        assert await wm.get_subject_by_name("Fresh", "c1") is not None

    @pytest.mark.asyncio
    async def test_requires_a_scoring_frame(self, wm) -> None:
        await wm.add_subject(name="Rival", company_id="c1", url="https://r.example")
        res = await (await self._tool(wm)).execute(
            {"subject": "Rival", "company_id": "c1", "save": False}
        )
        assert not res.success and "seed" in res.error

    @pytest.mark.asyncio
    async def test_unreadable_site_fails_loudly_without_inventing(
        self, wm, monkeypatch
    ) -> None:
        import core.watch_observe as wo

        async def dead(start_url, **kw):
            return [{"url": start_url, "text": "", "error": "HTTP 403", "method": "http"}]

        monkeypatch.setattr(wo, "collect_pages", dead)
        await wm.add_subject(name="Rival", company_id="c1", url="https://r.example")
        await wm.upsert_dimension(name="Promos", company_id="c1", weight_pct=100)
        res = await (await self._tool(wm)).execute(
            {"subject": "Rival", "company_id": "c1", "save": False}
        )
        assert not res.success and "could not read" in res.error
        assert await wm.list_evidence("c1") == []
