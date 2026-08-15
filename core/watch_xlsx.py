"""XLSX rendering for the competitive-intelligence scorecard.

The client-facing deliverable. Four sheets so a reader can go from the headline
ranking all the way down to the source URL behind any single score:

    Scorecard  — brands x dimensions, weighted totals, both alternative views
    Weights    — the scoring frame: weight, cadence, sub-criteria
    Evidence   — the full register with provenance (the audit trail)
    Gaps       — what is unscored, thin, or overdue for a refresh

One rendering rule matters more than any formatting choice: **an unscored
dimension is a blank cell, never a zero.** A zero would read as "terrible" when
it means "not yet observed", which is exactly the misreading the whole design
exists to prevent. The Gaps sheet says so in words.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_HEADER_BG = "1F2937"  # slate-800
_HEADER_FG = "FFFFFF"
_SELF_BG = "FEF3C7"  # amber-100 — our own brands
_GAP_BG = "F3F4F6"  # gray-100 — no evidence
_BORDER = "D1D5DB"


def _style_header(ws: Any, row: int, ncols: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=_HEADER_FG, size=10)
        cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )


def _autosize(ws: Any, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def render_scorecard_xlsx(
    card: dict[str, Any],
    *,
    dimensions: list[Any],
    evidence: list[Any],
    staleness: list[dict[str, Any]],
    path: str | Path,
    title: str = "Competitive Scorecard",
) -> str:
    """Write the four-sheet workbook. Returns the path written."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Scorecard ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Scorecard"
    dim_names = [d.name for d in dimensions]

    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(
        [
            f"Generated {card.get('generated_at', '')} · scores 1-5 "
            "· blank = no evidence yet (never a low score)"
        ]
    )
    ws["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws.append([])

    header = (
        ["#", "Brand", "Group"]
        + dim_names
        + [
            "Overall",
            "Customer proposition",
            "Transition priority",
            "Coverage %",
            "Confidence",
            "Scored weight %",
            "Status",
        ]
    )
    ws.append(header)
    _style_header(ws, 4, len(header))
    ws.freeze_panes = "D5"

    for r in card.get("rows", []):
        overall = r.get("overall", {})
        views = r.get("views", {})
        row = [
            r.get("rank") or "—",
            r.get("name", ""),
            r.get("group", ""),
        ]
        for dn in dim_names:
            d = r.get("dimensions", {}).get(dn, {})
            # Blank, not zero — the entire point.
            row.append(d.get("score") if d.get("score") is not None else None)
        row += [
            overall.get("normalized_pct"),
            views.get("customer_proposition", {}).get("normalized_pct"),
            views.get("transition_priority", {}).get("normalized_pct"),
            overall.get("coverage_pct"),
            overall.get("confidence", ""),
            overall.get("scored_weight_pct"),
            # Spelled out in the sheet itself. A client filters and sorts
            # this file; a caveat that lives only in the notes row travels
            # nowhere once a column is sorted or a row is copied out.
            (
                "Provisional — not ranked"
                if r.get("provisional") and overall.get("normalized_pct") is not None
                else (
                    "No evidence" if overall.get("normalized_pct") is None else "Ranked"
                )
            ),
        ]
        ws.append(row)
        if r.get("is_self"):
            for c in range(1, len(header) + 1):
                ws.cell(row=ws.max_row, column=c).fill = PatternFill(
                    "solid", fgColor=_SELF_BG
                )
        # Shade the dimension cells that have no score.
        for i, dn in enumerate(dim_names):
            d = r.get("dimensions", {}).get(dn, {})
            if d.get("score") is None:
                ws.cell(row=ws.max_row, column=4 + i).fill = PatternFill(
                    "solid", fgColor=_GAP_BG
                )

    _autosize(ws, {"A": 5, "B": 24, "C": 18})
    for i in range(len(dim_names)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14
    for i in range(6):
        ws.column_dimensions[get_column_letter(4 + len(dim_names) + i)].width = 16

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1).value = (
        "Overall / view figures are normalised to the weight actually scored, so "
        "brands with different evidence coverage remain comparable. A blank "
        "dimension means no evidence has been collected — it is NOT a low score. "
        "Read Coverage % and the Gaps sheet alongside every ranking."
    )
    ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="6B7280")
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)

    # ── Sheet 2: Weights ────────────────────────────────────────────
    ws2 = wb.create_sheet("Weights")
    hdr2 = [
        "Dimension",
        "Weight %",
        "Customer proposition %",
        "Transition priority %",
        "Refresh cadence",
        "Sub-criteria (weight %)",
        "Description",
    ]
    ws2.append(hdr2)
    _style_header(ws2, 1, len(hdr2))
    ws2.freeze_panes = "A2"
    for d in dimensions:
        subs = "; ".join(
            f"{s.get('name')} ({s.get('weight_pct')}%)" for s in (d.subcriteria or [])
        )
        ws2.append(
            [
                d.name,
                d.weight_pct,
                d.view_weights.get("customer_proposition"),
                d.view_weights.get("transition_priority"),
                d.refresh_cadence,
                subs,
                d.description,
            ]
        )
    total = round(sum(float(d.weight_pct) for d in dimensions), 2)
    ws2.append([])
    ws2.append(["TOTAL", total, "", "", "", "", "must equal 100"])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=2).font = Font(
        bold=True, color="B91C1C" if abs(total - 100) > 0.01 else "047857"
    )
    _autosize(
        ws2,
        {"A": 34, "B": 10, "C": 20, "D": 20, "E": 15, "F": 62, "G": 50},
    )

    # ── Sheet 3: Evidence register ──────────────────────────────────
    ws3 = wb.create_sheet("Evidence")
    hdr3 = [
        "Brand",
        "Dimension",
        "Sub-criterion",
        "Claim",
        "Value",
        "Source URL",
        "Source type",
        "Geo / state",
        "Customer state",
        "Observed at",
        "Confidence",
        "Collector",
        "Excerpt",
        "Screenshot",
        "Exit IP",
    ]
    ws3.append(hdr3)
    _style_header(ws3, 1, len(hdr3))
    ws3.freeze_panes = "A2"
    for e in evidence:
        ws3.append(
            [
                e.get("subject", ""),
                e.get("dimension", ""),
                e.get("subcriterion", ""),
                e.get("claim", ""),
                e.get("value_text", ""),
                e.get("source_url", ""),
                e.get("source_type", ""),
                e.get("geo_state", ""),
                e.get("customer_state", ""),
                e.get("observed_at", ""),
                e.get("confidence", ""),
                e.get("collector", ""),
                (e.get("excerpt", "") or "")[:500],
                e.get("screenshot_path", ""),
                e.get("exit_ip", ""),
            ]
        )
    _autosize(
        ws3,
        {
            "A": 20,
            "B": 30,
            "C": 24,
            "D": 52,
            "E": 18,
            "F": 44,
            "G": 13,
            "H": 12,
            "I": 15,
            "J": 22,
            "K": 12,
            "L": 11,
            "M": 60,
        },
    )

    # ── Sheet 4: Gaps ───────────────────────────────────────────────
    ws4 = wb.create_sheet("Gaps")
    ws4.append(["Evidence gaps and refresh status"])
    ws4["A1"].font = Font(bold=True, size=12)
    ws4.append(
        [
            "A gap is an absence of evidence, not a weakness. These entries "
            "explain every blank cell on the Scorecard sheet."
        ]
    )
    ws4["A2"].font = Font(italic=True, size=9, color="6B7280")
    ws4.append([])
    hdr4 = ["Brand", "Dimension", "Cadence", "Last observed", "Age (days)", "Status"]
    ws4.append(hdr4)
    _style_header(ws4, 4, len(hdr4))
    ws4.freeze_panes = "A5"
    for g in staleness:
        ws4.append(
            [
                g.get("subject", ""),
                g.get("dimension", ""),
                g.get("cadence", ""),
                g.get("last_observed") or "—",
                g.get("age_days") if g.get("age_days") is not None else "—",
                g.get("status", ""),
            ]
        )
    _autosize(ws4, {"A": 22, "B": 34, "C": 12, "D": 24, "E": 12, "F": 16})

    out = Path(path).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    logger.info("watch: scorecard workbook written to %s", out)
    return str(out)
